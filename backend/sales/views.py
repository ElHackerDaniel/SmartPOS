# sales/views.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Sum
from django.db.models.functions import TruncDay
from rest_framework import viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import Sale
from .serializers import SaleSerializer
from products.models import Product


class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all()
    serializer_class = SaleSerializer

    def perform_create(self, serializer):
        sale = serializer.save()
        product = sale.product
        if product.stock < sale.quantity:
            raise ValidationError("No hay suficiente stock")
        product.stock -= sale.quantity
        product.save()


@api_view(['GET'])
def sales_history(request):
    """Vista para obtener el historial de ventas"""
    sales = Sale.objects.all().order_by('-date')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        sales = sales.filter(date__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(date__lte=parse_date(end_date))
    
    serializer = SaleSerializer(sales, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def sales_by_day(request):
    """Ventas agrupadas por día"""
    start = request.GET.get("start")
    end = request.GET.get("end")

    sales = Sale.objects.all()

    if start:
        sales = sales.filter(date__date__gte=parse_date(start))
    if end:
        sales = sales.filter(date__date__lte=parse_date(end))

    sales = (
        sales
        .annotate(day=TruncDay('date'))
        .values('day')
        .annotate(total=Sum('sale_price'))
        .order_by('day')
    )

    data = [
        {
            "day": s["day"].strftime("%Y-%m-%d"),
            "sales": float(s["total"])
        }
        for s in sales
    ]

    return Response(data)


# 🔥 Endpoint para exportar ventas a Excel (VERSIÓN ÚNICA Y MEJORADA)
@api_view(['GET'])
def export_sales_excel(request):
    """Exportar ventas a Excel con formato profesional"""
    sales = Sale.objects.all().select_related('product').order_by('-date')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Estilos para encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ["ID", "Producto", "Cantidad", "Precio Unitario", "Total", "Fecha"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, sale in enumerate(sales, 2):
        ws.cell(row=row, column=1, value=sale.id)
        ws.cell(row=row, column=2, value=sale.product.name)
        ws.cell(row=row, column=3, value=sale.quantity)
        ws.cell(row=row, column=4, value=float(sale.sale_price))
        ws.cell(row=row, column=5, value=float(sale.sale_price) * sale.quantity)
        ws.cell(row=row, column=6, value=sale.date.strftime("%Y-%m-%d %H:%M"))
    
    # Ajustar ancho de columnas
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F']
    for col in column_letters:
        ws.column_dimensions[col].width = 18
    
    # Crear respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=ventas_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response